import openpyxl
from openpyxl.utils import get_column_letter

# Load your Excel file
file_path = "Grade Data Pulls (1).xlsx"
workbook = openpyxl.load_workbook(filename=file_path)

# Load the grade conversion table from the "Grade Convert" worksheet
grade_convert_sheet = workbook["Grade Convert"]
grade_conversion = {}

# Build the conversion dictionary (column B to G)
for row in grade_convert_sheet.iter_rows(min_row=2, min_col=2, max_col=7):
    letter = row[0].value
    value = row[5].value
    if letter is not None and value is not None:
        grade_conversion[str(letter).strip().upper()] = value

# Define the range N to T (columns 14 to 20)
columns_to_update = [get_column_letter(col) for col in range(14, 21)]

# Process all sheets except "Grade Convert"
for sheet_name in workbook.sheetnames:
    if sheet_name != "Grade Convert":
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2):
            for col_letter in columns_to_update:
                cell = sheet[f"{col_letter}{row[0].row}"]
                if isinstance(cell.value, str):
                    clean_val = cell.value.strip().upper()
                    if clean_val in grade_conversion:
                        cell.value = grade_conversion[clean_val]

# Save the updated file
output_path = "Grade_Data_Pulls_Converted.xlsx"
workbook.save(output_path)
print(f"✅ Conversion complete. File saved as: {output_path}")
